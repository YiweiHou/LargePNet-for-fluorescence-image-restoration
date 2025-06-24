import torch.nn as nn
import torch.nn.init as init
import os
from tqdm import tqdm
import torch
import numpy as np
import torch.utils.data as Data
from Utils.loss import *
from torch.cuda.amp import GradScaler, autocast
from Utils.ssim3d import *
from Utils.DataAug3D import *
import math
import torch.nn.functional as F
import pandas as pd


def rescale(restored, gt):
    '''Affine rescaling h'''

    batch_size = restored.size(0)
    restored_flat = restored.view(batch_size, -1)
    gt_flat = gt.view(batch_size, -1)

    mean_restored = restored_flat.mean()
    mean_gt = gt_flat.mean()
    cov_restored_gt = torch.mean((restored - mean_restored) * (gt - mean_gt))
    var_restored = torch.mean((restored - mean_restored) ** 2)

    a = cov_restored_gt / var_restored
    b = mean_gt - a * mean_restored

    return a * restored + b

def weights_init(m):
    if isinstance(m, nn.Linear):
        init.xavier_uniform_(m.weight)
        if m.bias is not None:
            init.zeros_(m.bias)


def weights_init_kaiming(m):
    classname = m.__class__.__name__
    if classname.find('Conv') != -1:
        nn.init.kaiming_normal(m.weight.data, a=0, mode='fan_in')
    elif classname.find('Linear') != -1:
        nn.init.kaiming_normal(m.weight.data, a=0, mode='fan_in')
    elif classname.find('BatchNorm') != -1:
        # nn.init.uniform(m.weight.data, 1.0, 0.02)
        m.weight.data.normal_(mean=0, std=math.sqrt(2. / 9. / 64.)).clamp_(-0.025, 0.025)
        nn.init.constant(m.bias.data, 0.0)


def TrainerAug(model, train_data, val_data, save_path, opt):
    if not os.path.exists(save_path):
        os.mkdir(save_path)
    batchsize = opt.batchsize
    scaler = GradScaler()
    MAE_weight = opt.MAE_weight
    SSIM_weight = opt.SSIM_weight
    LR = opt.LR
    epoch_num = opt.epoch_num
    epoch_critic = opt.epoch_critic

    train_loader = Data.DataLoader(
        dataset=train_data,
        batch_size=opt.batchsize,
        shuffle=True,
        num_workers=1,
    )

    val_loader = Data.DataLoader(
        dataset=val_data,
        batch_size=1,
        shuffle=True,
        num_workers=1,
    )
    if opt.useinit == 1:
        model.apply(weights_init)
        # model.apply(weights_init_kaiming)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = model.to(device)
    sigma_params = [(name, param) for name, param in model.named_parameters() if 'sigma' in name]

    sigma_param_names = {name for name, _ in sigma_params}

    other_params = [param for name, param in model.named_parameters() if name not in sigma_param_names]

    param_groups = [{'params': other_params, 'lr': LR},
                    {'params': [param for _, param in sigma_params], 'lr': LR / 1000}]

    optimizer = torch.optim.Adam(param_groups)
    criterion1 = nn.MSELoss()
    criterion2 = nn.L1Loss()
    criterion3 = SSIM3D()

    MSE_vals = []
    MAE_vals = []
    SSIM_vals = []
    MSE_trains = []
    MAE_trains = []
    best_val_loss = float('inf')

    for epoch in range(epoch_num):
        print(epoch)
        train_loss_epoch = 0
        val_loss_epoch = 0

        MSE_val_epoch = 0
        MAE_val_epoch = 0
        MSE_train_epoch = 0
        MAE_train_epoch = 0
        SSIM_val_epoch = 0
        train_num = 0
        val_num = 0
        if opt.StepLR == 1:
            for ii in range(len(opt.Epochcritic_Point)):
                if epoch > opt.Epochcritic_Point[ii] - 1:
                    param_groups = [{'params': other_params, 'lr': LR / opt.Epochcritic_Decay[ii]},
                                    {'params': [param for _, param in sigma_params],
                                     'lr': LR / opt.Epochcritic_Decay[ii] / 1000}]
                    optimizer = torch.optim.Adam(param_groups)
        if opt.ExpoLR == 1:
            param_groups = [{'params': other_params, 'lr': LR * (0.9 ** epoch)},
                            {'params': [param for _, param in sigma_params], 'lr': LR / 1000 * (0.9 ** epoch)}]
            optimizer = torch.optim.Adam(param_groups)

        train_loader = tqdm(train_loader, desc="Training epoch", total=len(train_loader))
        for step, (raw, gt) in enumerate(train_loader):
            for i in range(math.floor(opt.minibatchnum / opt.img_num)):
                b_x, b_y = DataAug(np.array(raw[0, ...]).astype(np.float32), np.array(gt[0, ...]).astype(np.float32),
                                   opt.patchsize)
                b_x = np.transpose(b_x, (2, 0, 1))
                b_y = np.transpose(b_y, (2, 0, 1))

                b_x = (b_x - b_x.min()) / (b_x.max() - b_x.min())
                b_y = (b_y - b_y.min()) / (b_y.max() - b_y.min())
                b_x = b_x.copy()
                b_y = b_y.copy()
                b_x = torch.tensor(b_x, dtype=torch.float16).unsqueeze(0).unsqueeze(0)
                b_y = torch.tensor(b_y, dtype=torch.float16).unsqueeze(0).unsqueeze(0)
                b_x = b_x.to(device)
                b_y = b_y.to(device)
                optimizer.zero_grad()
                model.train()
                with autocast():
                    output = model(b_x)
                    MSEtrain = criterion1(output, b_y)
                    MAEtrain = criterion2(output, b_y)
                    # SSIMtrain = criterion3(output,b_y)

                    loss = MSEtrain + MAE_weight * MAEtrain  # + SSIM_weight * (1-SSIMtrain)

                scaler.scale(loss).backward()

                scaler.step(optimizer)
                scaler.update()
                # optimizer.zero_grad()
                # loss.backward()
                # optimizer.step()
                train_loader.set_postfix_str(f"Epoch: {epoch + 1}, Loss: {loss.item()}")
                MSE_train_epoch += MSEtrain * b_x.size(0)
                MAE_train_epoch += MAEtrain * b_x.size(0)

                MSE_train_epoch += MSEtrain.cpu().detach() * batchsize
                MAE_train_epoch += MAEtrain.cpu().detach() * batchsize

                train_num = train_num + batchsize
                # del b_x, b_y, output, loss
                # torch.cuda.empty_cache()

        MSE_train = MSE_train_epoch.cpu().detach() / train_num
        MAE_train = MAE_train_epoch.cpu().detach() / train_num

        MSE_trains.append(MSE_train)
        MAE_trains.append(np.array(MAE_train))

        train_loss = MSE_train + MAE_weight * MAE_train

        with torch.no_grad():
            for step, (b_x, b_y) in enumerate(val_loader):

                b_x = (b_x - b_x.min()) / (b_x.max() - b_x.min())
                b_y = (b_y - b_y.min()) / (b_y.max() - b_y.min())

                b_x = b_x.to(device)
                b_y = b_y.to(device)

                model.eval()
                model = model.to('cpu')
                with autocast():
                    model = model.to(device)
                    # b_x = b_x / b_x.max()
                    b_y = b_y / b_y.max()
                    output = model(b_x)
                    #output = rescale(output, b_y)
                    rmse = torch.sqrt(torch.mean((output - b_y) ** 2))
                    MSEval = 20 * torch.log10(1 / rmse)
                    MAEval = criterion2(output, b_y)
                    SSIMval = criterion3(output, b_y)

                MSE_val_epoch += MSEval.cpu() * 1
                MAE_val_epoch += MAEval.cpu() * 1
                SSIM_val_epoch += SSIMval.cpu() * 1

                val_num = val_num + b_x.size(0)
                del MSEval, MAEval, SSIMval

            MSE_val = MSE_val_epoch / val_num
            MAE_val = MAE_val_epoch / val_num
            SSIM_val = SSIM_val_epoch / val_num
            val_loss = 50 - MSE_val + MAE_weight * MAE_val

        MSE_vals.append(MSE_val)
        MAE_vals.append(MAE_val)
        SSIM_vals.append(SSIM_val)
        df = pd.DataFrame({
            'Epoch': [epoch + 1],
            'MSE_Train': MSE_trains[-1],
            'MSE_Val': MSE_vals[-1],
            'MAE_Val': MAE_vals[-1]
        })

        file_path = save_path + '\\' + 'training_results.xlsx'

        if epoch == 0:
            df.to_csv(file_path, index=False, encoding='utf-8')
        else:
            df.to_csv(file_path, mode='a', header=False, index=False, encoding='utf-8')

        if val_loss < best_val_loss:
            best_val_loss = val_loss
            torch.save(model.state_dict(), os.path.join(save_path, 'best_model.pth'))
        path = os.path.join(save_path, 'final_network.pth')
        torch.save(model.state_dict(), path)
        print("Epoch: {}, Train Loss: {:.4f}, Train MSE: {:.4f}, Train MAE: {:.4f}".format(epoch, train_loss, MSE_train,
                                                                                           MAE_train))
        print("Epoch: {}, Val Loss: {:.4f}, Val MSE: {:.4f}, Val MAE: {:.4f}, Val SSIM: {:.4f}".format(epoch, val_loss,
                                                                                                       MSE_val, MAE_val,
                                                                                                       SSIM_val))

        torch.cuda.empty_cache()

    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MSE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'MSE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(SSIM_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'SSIM_vals.xlsx')