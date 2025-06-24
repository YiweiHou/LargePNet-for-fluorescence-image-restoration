import torch.nn as nn
import torch.nn.init as init
import os
from tqdm import tqdm
import torch
import numpy as np
import torch.utils.data as Data
from Utils.loss import *
from torch.cuda.amp import GradScaler, autocast
import math
import tifffile


def weights_init(m):
    if isinstance(m, nn.Linear):
        init.xavier_uniform_(m.weight)
        if m.bias is not None:
            init.zeros_(m.bias)


def weights_init_kaiming(m):
    classname = m.__class__.__name__
    if classname.find('Conv') != -1:
        nn.init.kaiming_normal(m.weight.data, a=0, mode='fan_in')
    elif classname.find('Linear') != -1:
        nn.init.kaiming_normal(m.weight.data, a=0, mode='fan_in')
    elif classname.find('BatchNorm') != -1:
        # nn.init.uniform(m.weight.data, 1.0, 0.02)
        m.weight.data.normal_(mean=0, std=math.sqrt(2. / 9. / 64.)).clamp_(-0.025, 0.025)
        nn.init.constant(m.bias.data, 0.0)


def Show_img(model, img_dir, save_dir, epoch_num):
    if len(img_dir) > 0:
        device = 'cuda'
        Endstr = '.tif'
        numstr = np.array2string(np.array(epoch_num + 1))
        with torch.no_grad():
            b_x = tifffile.imread(img_dir)
            b_x = np.array(b_x).astype(np.float32)
            b_x = (b_x - b_x.min()) / (b_x.max() - b_x.min())
            b_x = torch.tensor(b_x, dtype=torch.float16)
            b_x = b_x.to(device)
            b_x = b_x.unsqueeze(0).unsqueeze(0)

            model.eval()
            model = model.to(device)
            with autocast():
                output = model(b_x)

            output = output.to('cpu')
            output = output.detach()

            imde = np.transpose(output.data.numpy(), (0, 2, 3, 1))
            imdeHybrid = imde[0, ...]
            imde = imdeHybrid
            imde = (2 ** 16 - 1) * imde
            # imde = (2**16-1)*(imde-imde.min())/(imde.max()-imde.min())
            imde = imde.astype(np.uint16)
            tifffile.imwrite(save_dir + '\\' + numstr + Endstr, imde)
    return True


class EdgeAwareLoss(nn.Module):
    def __init__(self, edge_weight=0.05):
        super().__init__()
        self.edge_weight = edge_weight

        kernel = torch.tensor([[0, 1, 0],
                               [1, -4, 1],
                               [0, 1, 0]], dtype=torch.float32)
        self.edge_conv = nn.Conv2d(1, 1, kernel_size=3, padding=0, bias=False)
        self.edge_conv.weight.data = kernel.view(1, 1, 3, 3)
        self.edge_conv.weight.requires_grad = False

        self.mse = nn.MSELoss()

    def forward(self, pred, target):
        c_loss = self.mse(pred, target)


        with torch.cuda.amp.autocast(enabled=False):
            pred_edge = self.edge_conv(pred.float())
            target_edge = self.edge_conv(target.float())
            e_loss = self.mse(pred_edge, target_edge)

        return c_loss + self.edge_weight * e_loss


def TrainerSTED(model, train_data, val_data, save_path, opt):
    if not os.path.exists(save_path):
        os.mkdir(save_path)
    batchsize = opt.batchsize
    scaler = GradScaler()
    MAE_weight = opt.MAE_weight
    SSIM_weight = opt.SSIM_weight
    LR = opt.LR

    epoch_num = opt.epoch_num
    epoch_critic = opt.epoch_critic

    train_loader = Data.DataLoader(
        dataset=train_data,
        batch_size=opt.batchsize,
        shuffle=True,
        num_workers=1,
    )

    val_loader = Data.DataLoader(
        dataset=val_data,
        batch_size=1,
        shuffle=True,
        num_workers=1,
    )
    if opt.useinit == 1:
        model.apply(weights_init)
    elif opt.useinit == 2:
        model.apply(weights_init_kaiming)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = model.to(device)

    optimizer = torch.optim.Adam(model.parameters(), lr=LR)
    criterion1 = EdgeAwareLoss().to('cuda')
    criterion2 = nn.L1Loss()
    criterion3 = SSIM()

    MSE_vals = []
    MAE_vals = []
    SSIM_vals = []
    MSE_trains = []
    MAE_trains = []
    best_val_loss = float('inf')

    for epoch in range(epoch_num):
        epoch_indi = epoch + opt.pre_epoch
        print(epoch)
        train_loss_epoch = 0
        val_loss_epoch = 0

        MSE_val_epoch = 0
        MAE_val_epoch = 0
        MSE_train_epoch = 0
        MAE_train_epoch = 0
        SSIM_val_epoch = 0
        train_num = 0
        val_num = 0
        if opt.StepLR == 1:
            if epoch > epoch_critic:

                optimizer = torch.optim.Adam(model.parameters(), lr=LR / 2)

        train_loader = tqdm(train_loader, desc="Training epoch", total=len(train_loader))
        for step, (b_x, b_y) in enumerate(train_loader):
            b_x = b_x.to(device)
            b_y = b_y.to(device)
            optimizer.zero_grad()
            model.train()
            with autocast():
                output = model(b_x)
                # output = (output-output.min())/(output.max()-output.min())
                MSEtrain = criterion1(output, b_y)
                MAEtrain = criterion2(output, b_y)
                SSIMtrain = criterion3(output, b_y)

                loss = MSEtrain  # + SSIM_weight * (1-SSIMtrain)

            scaler.scale(loss).backward()

            # 应用缩放后的梯度
            scaler.step(optimizer)
            scaler.update()

            # optimizer.zero_grad()
            # loss.backward()
            # optimizer.step()
            train_loader.set_postfix_str(f"Epoch: {epoch + 1}, Loss: {loss.item()}")
            MSE_train_epoch += MSEtrain * b_x.size(0)
            MAE_train_epoch += MAEtrain * b_x.size(0)

            MSE_train_epoch += MSEtrain.cpu().detach() * batchsize
            MAE_train_epoch += MAEtrain.cpu().detach() * batchsize

            train_num = train_num + batchsize
            # del b_x, b_y, output, loss
            # torch.cuda.empty_cache()

        MSE_train = MSE_train_epoch.cpu().detach() / train_num
        MAE_train = MAE_train_epoch.cpu().detach() / train_num

        MSE_trains.append(MSE_train)
        MAE_trains.append(np.array(MAE_train))

        train_loss = MSE_train

        with torch.no_grad():
            for step, (b_x, b_y) in enumerate(val_loader):
                b_x = b_x.to(device)
                b_y = b_y.to(device)

                model.eval()
                model = model.to('cpu')
                with autocast():
                    model = model.to(device)
                    output = model(b_x)
                    # output = (output-output.min())/(output.max()-output.min())
                    MSEval = criterion1(output, b_y)
                    MAEval = criterion2(output, b_y)
                    SSIMval = criterion3(output, b_y)

                MSE_val_epoch += MSEval.cpu() * 1
                MAE_val_epoch += MAEval.cpu() * 1
                SSIM_val_epoch += SSIMval.cpu() * 1

                val_num = val_num + b_x.size(0)
                del MSEval, MAEval, SSIMval

            MSE_val = MSE_val_epoch / val_num
            MAE_val = MAE_val_epoch / val_num
            SSIM_val = SSIM_val_epoch / val_num
            val_loss = MSE_val

        MSE_vals.append(MSE_val)
        MAE_vals.append(MAE_val)
        SSIM_vals.append(SSIM_val)
        Show_img(model, opt.instanceimage, save_path, epoch_indi)

        if val_loss < best_val_loss:
            best_val_loss = val_loss
            torch.save(model.state_dict(), os.path.join(save_path, 'best_model.pth'))
        path = os.path.join(save_path, 'final_network.pth')
        torch.save(model.state_dict(), path)
        print("Epoch: {}, Train Loss: {:.6f}, Train MSE: {:.6f}, Train MAE: {:.6f}".format(epoch, train_loss, MSE_train,
                                                                                           MAE_train))
        print("Epoch: {}, Val Loss: {:.6f}, Val MSE: {:.6f}, Val MAE: {:.6f}, Val SSIM: {:.6f}".format(epoch, val_loss,
                                                                                                       MSE_val, MAE_val,
                                                                                                       SSIM_val))

        torch.cuda.empty_cache()
        # del train_loss_epoch, val_loss_epoch, MSE_val_epoch, MAE_val_epoch, MSE_train_epoch, MAE_train_epoch, train_num, val_num, b_x, b_y

    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MSE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'MSE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(SSIM_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'SSIM_vals.xlsx')