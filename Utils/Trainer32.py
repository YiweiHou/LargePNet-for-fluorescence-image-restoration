import torch.nn as nn
import torch.nn.init as init
import os
from tqdm import tqdm
import torch
import numpy as np
import torch.utils.data as Data
from Utils.loss import *
from torch.cuda.amp import GradScaler, autocast
import tifffile


def weights_init(m):
    if isinstance(m, nn.Linear):
        init.xavier_uniform_(m.weight)
        if m.bias is not None:
            init.zeros_(m.bias)


def weights_init_kaiming(m):
    classname = m.__class__.__name__
    if classname.find('Conv') != -1:
        nn.init.kaiming_normal(m.weight.data, a=0, mode='fan_in')
    elif classname.find('Linear') != -1:
        nn.init.kaiming_normal(m.weight.data, a=0, mode='fan_in')
    elif classname.find('BatchNorm') != -1:
        # nn.init.uniform(m.weight.data, 1.0, 0.02)
        m.weight.data.normal_(mean=0, std=math.sqrt(2. / 9. / 64.)).clamp_(-0.025, 0.025)
        nn.init.constant(m.bias.data, 0.0)


def Val_step(val_data, model, criterion2, criterion3):
    device = 'cuda'
    val_num = 0
    val_loss_epoch = 0
    MSE_val_epoch = 0
    MAE_val_epoch = 0
    PSNR_val_epoch = 0
    SSIM_val_epoch = 0
    with torch.no_grad():
        for i in range(len(val_data)):
            b_x, b_y = val_data[i]
            b_x = b_x.to(device)
            b_y = b_y.to(device)
            b_x = b_x.unsqueeze(0)
            b_y = b_y.unsqueeze(0)

            model.eval()
            model = model.to(device)
            output = model(b_x)

            MSEval = torch.mean((output - b_y) ** 2)
            PSNRval = 10 * torch.log10(1 / MSEval)
            MAEval = criterion2(output, b_y)
            SSIMval = criterion3(output, b_y)

            MSE_val_epoch += MSEval.cpu() * 1
            MAE_val_epoch += MAEval.cpu() * 1
            SSIM_val_epoch += SSIMval.cpu() * 1
            PSNR_val_epoch += PSNRval.cpu() * 1

            val_num = val_num + b_x.size(0)
            del MSEval, MAEval, SSIMval

        MSE_val = MSE_val_epoch / val_num
        MAE_val = MAE_val_epoch / val_num
        SSIM_val = SSIM_val_epoch / val_num
        PSNR_val = PSNR_val_epoch / val_num
        val_loss = MSE_val
    return MSE_val, MAE_val, SSIM_val, PSNR_val, val_loss



def Trainer32(model, train_data, val_data, save_path, opt):
    if not os.path.exists(save_path):
        os.mkdir(save_path)
    batchsize = opt.batchsize
    scaler = GradScaler()
    MAE_weight = opt.MAE_weight
    SSIM_weight = opt.SSIM_weight
    LR = opt.LR
    epoch_num = opt.epoch_num
    epoch_critic = opt.epoch_critic

    train_loader = Data.DataLoader(
        dataset=train_data,
        batch_size=opt.batchsize,
        shuffle=True,
        num_workers=1,
    )

    # val_loader = Data.DataLoader(
    #     dataset=val_data,
    #     batch_size=1,
    #     shuffle=True,
    #     num_workers=1,
    # )

    if opt.useinit == 1:
        model.apply(weights_init)
    elif opt.useinit == 2:
        model.apply(weights_init_kaiming)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = model.to(device)

    optimizer = torch.optim.Adam(model.parameters(), lr= opt.LR)
    criterion1 = nn.MSELoss()
    criterion2 = nn.L1Loss()
    criterion3 = SSIM()

    MSE_vals = []
    MAE_vals = []
    SSIM_vals = []
    MSE_trains = []
    MAE_trains = []
    best_val_loss = float('inf')
    best_val_PSNR = 0
    for epoch in range(epoch_num):
        print(epoch)
        # Refresh epoch recording variables
        train_loss_epoch = 0
        MSE_train_epoch = 0
        MAE_train_epoch = 0
        train_num = 0
        val_num = 0

        if epoch > epoch_critic:
            optimizer = torch.optim.Adam(model.parameters(), lr= opt.LR/2)

        train_loader = tqdm(train_loader, desc="Training epoch", total=len(train_loader))
        for step, (b_x, b_y) in enumerate(train_loader):
            b_x = b_x.to(device)
            b_y = b_y.to(device)
            if opt.use_norm == 1:
                b_x = (b_x - b_x.min()) / (b_x.max() - b_x.min())
                b_y = (b_y - b_y.min()) / (b_y.max() - b_y.min())
            optimizer.zero_grad()
            model.train()

            output = model(b_x)

            MSEtrain = criterion1(output, b_y)
            MAEtrain = criterion2(output, b_y)
            # SSIMtrain = criterion3(output,b_y)

            loss = MSEtrain + MAE_weight * MAEtrain  # + SSIM_weight * (1-SSIMtrain)
            # scaler.scale(loss).backward()
            # scaler.step(optimizer)
            # scaler.update()

            loss.backward()
            optimizer.step()
            train_loader.set_postfix_str(f"Epoch: {epoch + 1}, Loss: {scaler.scale(loss).item() / 65535}")
            MSE_train_epoch += MSEtrain * b_x.size(0)
            MAE_train_epoch += MAEtrain * b_x.size(0)

            MSE_train_epoch += MSEtrain.cpu().detach() * batchsize
            MAE_train_epoch += MAEtrain.cpu().detach() * batchsize

            train_num = train_num + batchsize
            if np.isin(step, opt.innerpoint):

                MSE_val, MAE_val, SSIM_val, PSNR_val, val_loss = Val_step(val_data, model, criterion2, criterion3)

                if PSNR_val > best_val_PSNR:
                    best_val_PSNR = PSNR_val
                    print("Best ValPSNR: {:.6f}".format(PSNR_val))
                    torch.save(model.state_dict(), os.path.join(save_path, 'best_model.pth'))

        MSE_train = MSE_train_epoch.cpu().detach() / train_num
        MAE_train = MAE_train_epoch.cpu().detach() / train_num

        MSE_trains.append(MSE_train)
        MAE_trains.append(np.array(MAE_train))

        train_loss = MSE_train + MAE_weight * MAE_train

        # Validation step
        if opt.val == 1:
            MSE_val, MAE_val, SSIM_val, PSNR_val, val_loss = Val_step(val_data, model, criterion2, criterion3)

            MSE_vals.append(MSE_val)
            MAE_vals.append(MAE_val)
            SSIM_vals.append(SSIM_val)


        if PSNR_val > best_val_PSNR:
            best_val_PSNR = PSNR_val
            print("Best ValPSNR: {:.6f}".format(PSNR_val))
            torch.save(model.state_dict(), os.path.join(save_path, 'best_model.pth'))

        path = os.path.join(save_path, 'final_network.pth')
        torch.save(model.state_dict(), path)
        print("Epoch: {}, Train Loss: {:.6f}, Train MSE: {:.6f}, Train MAE: {:.6f}".format(epoch, train_loss, MSE_train,
                                                                                           MAE_train))
        print("Epoch: {}, Val Loss: {:.6f}, Val MSE: {:.6f}, Val MAE: {:.6f}, Val SSIM: {:.6f}".format(epoch, val_loss,
                                                                                                       MSE_val, MAE_val,
                                                                                                       SSIM_val))

        torch.cuda.empty_cache()
        # del train_loss_epoch, val_loss_epoch, MSE_val_epoch, MAE_val_epoch, MSE_train_epoch, MAE_train_epoch, train_num, val_num, b_x, b_y
    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MSE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'MSE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(SSIM_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'SSIM_vals.xlsx')
    return model

