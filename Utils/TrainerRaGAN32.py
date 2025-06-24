import torch.nn as nn
import torch.nn.init as init
import os
from tqdm import tqdm
import torch
import numpy as np
import torch.utils.data as Data
from torch.cuda.amp import GradScaler, autocast
import torchvision.models as models
from Utils.GANLoss import *
from torchvision.models import vgg19, vgg16
from Utils.FastMSSSIM import *
import tifffile
from Utils.DataAug2D import *


def weights_init(m):
    if isinstance(m, nn.Linear):
        init.xavier_uniform_(m.weight)
        if m.bias is not None:
            init.zeros_(m.bias)


def Show_img(model, img_dir, save_dir, epoch_num):
    if len(img_dir) > 0:
        device = 'cuda'
        Endstr = '.tif'
        numstr = np.array2string(np.array(epoch_num + 1))
        with torch.no_grad():
            b_x = tifffile.imread(img_dir)
            b_x = np.array(b_x).astype(np.float32)
            b_x = (b_x - b_x.min()) / (b_x.max() - b_x.min())
            b_x = torch.tensor(b_x, dtype=torch.float16)
            b_x = b_x.to(device)
            b_x = b_x.unsqueeze(0).unsqueeze(0)

            model.eval()
            model = model.to(device)
            with autocast():
                output = model(b_x)

            output = output.to('cpu')
            output = output.detach()

            imde = np.transpose(output.data.numpy(), (0, 2, 3, 1))
            imdeHybrid = imde[0, ...]
            imde = imdeHybrid
            imde = (2 ** 16 - 1) * imde
            # imde = (2**16-1)*(imde-imde.min())/(imde.max()-imde.min())
            imde = imde.astype(np.uint16)
            imde[imde < 0] = 0
            tifffile.imwrite(save_dir + '\\' + numstr + Endstr, imde)
    return True


def Val_step(val_data, model, criterion2, criterion3):
    device = 'cuda'
    val_num = 0
    val_loss_epoch = 0
    MSE_val_epoch = 0
    MAE_val_epoch = 0
    PSNR_val_epoch = 0
    SSIM_val_epoch = 0

    with torch.no_grad():
        for i in range(len(val_data)):
            b_x, b_y = val_data[i]
            b_x = b_x.to(device)
            b_y = b_y.to(device)
            b_x = b_x.unsqueeze(0)
            b_y = b_y.unsqueeze(0)

            model.eval()
            model = model.to(device)

            output = model(b_x)

            MSEval = torch.mean((output - b_y) ** 2)
            PSNRval = 10 * torch.log10(1 / MSEval)
            MAEval = criterion2(output, b_y)
            SSIMval = criterion3(output, b_y)

            MSE_val_epoch += MSEval.cpu() * 1
            MAE_val_epoch += MAEval.cpu() * 1
            SSIM_val_epoch += SSIMval.cpu() * 1
            PSNR_val_epoch += PSNRval.cpu() * 1

            val_num = val_num + b_x.size(0)
            del MSEval, MAEval, SSIMval

        MSE_val = MSE_val_epoch / val_num
        MAE_val = MAE_val_epoch / val_num
        SSIM_val = SSIM_val_epoch / val_num
        PSNR_val = PSNR_val_epoch / val_num
        val_loss = MSE_val
    return MSE_val, MAE_val, SSIM_val, PSNR_val, val_loss


class PerceptualLoss(nn.Module):
    def __init__(self):
        super(PerceptualLoss, self).__init__()
        vgg = vgg16(pretrained=True)
        loss_network = nn.Sequential(*list(vgg.features)[:31]).eval()
        for param in loss_network.parameters():
            param.requires_grad = False
        self.loss_network = loss_network.to('cuda')
        self.mse_loss = nn.MSELoss()
        self.tv_loss = TVLoss()

    def forward(self, out_images, target_images):
        # 确保图像是三个通道
        if out_images.size(1) == 1:
            out_images = out_images.repeat(1, 3, 1, 1)
        if target_images.size(1) == 1:
            target_images = target_images.repeat(1, 3, 1, 1)

        # Perception Loss
        perception_loss = self.mse_loss(self.loss_network(out_images), self.loss_network(target_images))
        return perception_loss


def TrainerRaGAN32(netG, netD, train_data, val_data, save_path, opt):
    if not os.path.exists(save_path):
        os.mkdir(save_path)
    batchsize = opt.batchsize
    scaler = GradScaler()

    MAE_weight = opt.MAE_weight
    SSIM_weight = opt.SSIM_weight
    LR = opt.LR
    epoch_num = opt.epoch_num
    epoch_critic = opt.epoch_critic
    best_val_PSNR = 0
    generator_criterion = GeneratorLoss()
    train_loader = Data.DataLoader(
        dataset=train_data,
        batch_size=opt.batchsize,
        shuffle=True,
        num_workers=1,
    )

    val_loader = Data.DataLoader(
        dataset=val_data,
        batch_size=1,
        shuffle=True,
        num_workers=1,
    )
    if opt.useinit == 1:
        netG.apply(weights_init)
        netD.apply(weights_init)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    netG = netG.to(device)
    netD = netD.to(device)
    feature_extractor = PerceptualLoss()

    # Set feature extractor to inference mode
    # feature_extractor.eval()



    optimizerG = torch.optim.Adam(netG.parameters(), LR)
    optimizerD = torch.optim.Adam(netD.parameters(), LR / opt.D_decay)
    generator_criterion.cuda()
    results = {'d_loss': [], 'g_loss': [], 'd_score': [], 'g_score': [], 'psnr': [], 'ssim': []}

    criterion1 = nn.MSELoss()
    criterion2 = nn.L1Loss()
    criterion3 = MS_SSIM(channel=1)
    criterion_GAN = torch.nn.BCEWithLogitsLoss()
    # criterion_GAN = torch.nn.BCELoss()
    criterion_content = PerceptualLoss()

    MSE_vals = []
    MAE_vals = []
    SSIM_vals = []
    MSE_trains = []
    MAE_trains = []
    best_val_loss = float('inf')
    best_val_GAN_loss = float('inf')
    original_GAN = opt.loss_GAN
    for epoch in range(epoch_num):
        running_results = {'batch_sizes': 0, 'd_loss': 0, 'g_loss': 0, 'd_score': 0, 'g_score': 0}
        print(epoch)
        netG.train()
        netD.train()
        train_loss_epoch = 0
        val_loss_epoch = 0

        MSE_val_epoch = 0
        MAE_val_epoch = 0
        MSE_train_epoch = 0
        MAE_train_epoch = 0
        SSIM_val_epoch = 0
        train_num = 0
        val_num = 0

        if epoch > epoch_critic:

            optimizerG = torch.optim.Adam(netG.parameters(), lr=LR / 2)
            optimizerD = torch.optim.Adam(netD.parameters(), lr=LR / 2 / opt.D_decay)
        # if ( epoch + opt.Dinterval) % opt.Dinterval == 0:
        #     #opt.loss_GAN = original_GAN
        #     d_weight = 1
        # else:
        #     #opt.loss_GAN = 0
        #     d_weight = 0
        train_loader = tqdm(train_loader, desc="Training epoch", total=len(train_loader))
        for step, (b_x, b_y) in enumerate(train_loader):
            # b_x, b_y = DataAugKeep(b_x, b_y)
            valid = torch.ones((b_x.size(0), *netD.output_shape), requires_grad=False).to(device)
            fake = torch.zeros((b_x.size(0), *netD.output_shape), requires_grad=False).to(device)

            g_update_first = True
            batch_size = b_x.size(0)
            running_results['batch_sizes'] += batch_size
            b_x = b_x.to(device)
            b_y = b_y.to(device)
            if opt.use_norm == 1:
                b_x = (b_x - b_x.min()) / (b_x.max() - b_x.min())
                b_y = (b_y - b_y.min()) / (b_y.max() - b_y.min())
            # (1) Update G network: minimize 1-D(G(z)) + Perception Loss + Image Loss + TV Loss
            # (1-1) pixel loss
            optimizerG.zero_grad()
            fake_img = netG(b_x)
            MSEtrain = criterion1(fake_img, b_y)
            MAEtrain = criterion2(fake_img, b_y)
            SSIMtrain = 1 - criterion3(fake_img, b_y)
            pixel_loss = opt.MSE_weight * MSEtrain + opt.MAE_weight * MAEtrain  # +  opt.SSIM_weight * SSIMtrain

            # (1-2) content loss

            # Content loss

            # with torch.no_grad():
            #     gen_features = feature_extractor(fake_img)
            #     real_features = feature_extractor(b_y).detach()

            # with torch.no_grad():
            # loss_content = criterion_content(fake_img, b_y).detach()
            pred_real = netD(b_y).detach()
            pred_fake = netD(fake_img)
            loss_GAN = criterion_GAN(pred_fake - pred_real.mean(0, keepdim=True), valid)
            # loss_GAN = criterion_GAN(pred_fake, valid)
            # loss_GAN = 1 - torch.mean(pred_fake)
            loss_content = criterion_content(fake_img, b_y)
            loss_G = opt.loss_GAN * loss_GAN + pixel_loss + opt.loss_content * loss_content

            loss_G.backward()
            optimizerG.step()

            # (2) Update D network: maximize D(x)-1-D(G(z))
            # with autocast():
            pred_real = netD(b_y.float())
            pred_fake = netD(fake_img.detach().float())
            loss_real = criterion_GAN(pred_real - pred_fake.mean(0, keepdim=True), valid)
            loss_fake = criterion_GAN(pred_fake - pred_real.mean(0, keepdim=True), fake)
            # loss_real = criterion_GAN(pred_real, valid)
            # loss_fake = criterion_GAN(pred_fake, fake)
            # loss_real = torch.mean(pred_real)
            # loss_fake = torch.mean(pred_fake)
            loss_D = (loss_real + loss_fake) / 2
            optimizerD.zero_grad()
            loss_D.backward()
            optimizerD.step()
            # #if epoch > opt.warmupepoch-1:
            # scaler.scale(loss_D).backward()
            # scaler.step(optimizerD)
            # scaler.update()

            train_loader.set_postfix_str(f"Epoch: {epoch + 1}, Loss: {loss_G.item()}")
            if np.isin(step, opt.innerpoint):

                MSE_val, MAE_val, SSIM_val, PSNR_val, val_loss = Val_step(val_data, netG, criterion2, criterion3)


                if PSNR_val > best_val_PSNR:
                    best_val_PSNR = PSNR_val
                    print("Best ValPSNR: {:.6f}".format(PSNR_val))
                    torch.save(netG.state_dict(), os.path.join(save_path, 'best_model.pth'))
            # loss for current batch before optimization
            running_results['g_loss'] += loss_G.item() * batch_size
            running_results['d_loss'] += loss_D.item() * batch_size
            running_results['d_score'] += loss_real * batch_size
            running_results['g_score'] += loss_fake * batch_size

        # Validation step
        if opt.val == 1:

            MSE_val, MAE_val, SSIM_val, PSNR_val, val_loss = Val_step(val_data, netG, criterion2, criterion3)

            MSE_vals.append(MSE_val)
            MAE_vals.append(MAE_val)
            SSIM_vals.append(SSIM_val)
        Show_img(netG, opt.instanceimage, save_path, epoch)

        if epoch > opt.val_start:
            if PSNR_val > best_val_PSNR:
                best_val_PSNR = PSNR_val
                print("Best ValPSNR: {:.6f}".format(PSNR_val))
                torch.save(netG.state_dict(), os.path.join(save_path, 'best_model.pth'))

        path = os.path.join(save_path, 'final_networkG.pth')
        torch.save(netG.state_dict(), path)
        path = os.path.join(save_path, 'final_networkD.pth')
        torch.save(netD.state_dict(), path)
        print("Epoch: {}, Loss_D: {:.4f}, Loss_G: {:.4f}, D(x): {:.4f}, D(z): {:.4f}".format(epoch,
                                                                                             running_results['d_loss'] /
                                                                                             running_results[
                                                                                                 'batch_sizes'],
                                                                                             running_results['g_loss'] /
                                                                                             running_results[
                                                                                                 'batch_sizes'],
                                                                                             running_results[
                                                                                                 'd_score'] /
                                                                                             running_results[
                                                                                                 'batch_sizes'],
                                                                                             running_results[
                                                                                                 'g_score'] /
                                                                                             running_results[
                                                                                                 'batch_sizes']))
        print("Epoch: {}, Val Loss: {:.6f}, Val MSE: {:.6f}, Val MAE: {:.6f}, Val SSIM: {:.6f}".format(epoch, val_loss,
                                                                                                       MSE_val, MAE_val,
                                                                                                       SSIM_val))
        torch.cuda.empty_cache()
        # del train_loss_epoch, val_loss_epoch, MSE_val_epoch, MAE_val_epoch, MSE_train_epoch, MAE_train_epoch, train_num, val_num, b_x, b_y

    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MSE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MAE_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MAE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'MSE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(SSIM_vals)
    for row in a:
        sheet.append([row])
    workbook.save(save_path + '\\' + 'SSIM_vals.xlsx')